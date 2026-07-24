"""Export Excel du journal comptable — un onglet Synthèse + un onglet par type d'opération."""
import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

HEADER_FILL = PatternFill("solid", fgColor="2A1045")
HEADER_FONT = Font(color="E9CF8E", bold=True, size=10)
TOTAL_FILL = PatternFill("solid", fgColor="FBF6EE")
TOTAL_FONT = Font(bold=True, size=10)
THIN = Border(*[Side(style="thin", color="D9D2C4")] * 4)
EUR_FMT = '#,##0.00 "€"'
COLS = ["Date", "Libellé", "Pays", "Email", "Référence", "HT (EUR)", "TVA (EUR)", "TTC (EUR)"]


def _sheet_title(label: str) -> str:
    return re.sub(r"[\\/*?:\[\]]", "-", label)[:31]


def _write_header(ws):
    for i, c in enumerate(COLS, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, THIN
        cell.alignment = Alignment(horizontal="center")
    widths = [19, 62, 7, 30, 20, 13, 13, 13]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"


def _write_rows(ws, entries):
    for r, e in enumerate(entries, 2):
        vals = [e["date"][:19].replace("T", " "), e["label"], e["country"], e["email"], e["ref"],
                e["ht_cents"] / 100, e["vat_cents"] / 100, e["ttc_cents"] / 100]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = THIN
            if c >= 6:
                cell.number_format = EUR_FMT
    total_row = len(entries) + 2
    ws.cell(row=total_row, column=2, value=f"TOTAL ({len(entries)} opération(s))")
    for c, key in ((6, "ht_cents"), (7, "vat_cents"), (8, "ttc_cents")):
        cell = ws.cell(row=total_row, column=c, value=sum(e[key] for e in entries) / 100)
        cell.number_format = EUR_FMT
    for c in range(1, 9):
        cell = ws.cell(row=total_row, column=c)
        cell.fill, cell.font, cell.border = TOTAL_FILL, TOTAL_FONT, THIN


def build_accounting_workbook(entries: list, kind_labels: dict) -> bytes:
    """Classeur : onglet Synthèse (ventilation par type) + un onglet par type d'opération."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Synthèse"
    headers = ["Type d'opération", "Opérations", "HT (EUR)", "TVA (EUR)", "TTC (EUR)"]
    for i, c in enumerate(headers, 1):
        cell = ws.cell(row=1, column=i, value=c)
        cell.fill, cell.font, cell.border = HEADER_FILL, HEADER_FONT, THIN
    for i, w in enumerate([34, 12, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = w
    by_type = {}
    for e in entries:
        by_type.setdefault(e["type"], []).append(e)
    row = 2
    for t, group in sorted(by_type.items(), key=lambda kv: kind_labels.get(kv[0], kv[0])):
        vals = [kind_labels.get(t, t), len(group), sum(e["ht_cents"] for e in group) / 100,
                sum(e["vat_cents"] for e in group) / 100, sum(e["ttc_cents"] for e in group) / 100]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(row=row, column=c, value=v)
            cell.border = THIN
            if c >= 3:
                cell.number_format = EUR_FMT
        row += 1
    for c, v in enumerate(["TOTAL GÉNÉRAL", len(entries), sum(e["ht_cents"] for e in entries) / 100,
                           sum(e["vat_cents"] for e in entries) / 100,
                           sum(e["ttc_cents"] for e in entries) / 100], 1):
        cell = ws.cell(row=row, column=c, value=v)
        cell.fill, cell.font, cell.border = TOTAL_FILL, TOTAL_FONT, THIN
        if c >= 3:
            cell.number_format = EUR_FMT
    for t, group in sorted(by_type.items(), key=lambda kv: kind_labels.get(kv[0], kv[0])):
        sheet = wb.create_sheet(_sheet_title(kind_labels.get(t, t)))
        _write_header(sheet)
        _write_rows(sheet, group)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
